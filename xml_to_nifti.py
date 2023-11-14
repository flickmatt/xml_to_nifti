import xml.etree.ElementTree as ET
import openpyxl
import SimpleITK as sitk
import pandas as pd

def parse_xml(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()

    coordinates = []
    skip_coordinates = False

    for element in root.iter():
        if skip_coordinates:
            skip_coordinates = False
            continue

        if element.text is not None and '\\' in element.text:
            if element.text.strip() == 'CLOSED_PLANAR\n1':
                skip_coordinates = True
                continue
            
            coords = element.text.split('\\')
            if '0/255/0' not in coords:
                coordinates.append(coords)

    return coordinates

def write_coordinates_to_excel(coordinates, excel_file_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set column headers as "x", "y", "z"
    headers = ["x", "y", "z"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    for row, coord_set in enumerate(coordinates, start=2):  # Start from row 2 for data
        # Check if the coordinates are not (0, 255, 0)
        if coord_set != ['0', '255', '0']:
            for col, value in enumerate(coord_set, start=1):
                ws.cell(row=row, column=col, value=float(value))

    # Remove empty rows
    for row in reversed(range(1, ws.max_row + 1)):
        if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
            ws.delete_rows(row)

    wb.save(excel_file_path)

if __name__ == "__main__":
    # Script 1: Parse XML and write to Excel
    xml_file_path = "_STEALTH_ROIs_.xml"
    excel_file_path = "_STEALTH_ROIs_.xlsx"

    parsed_coordinates = parse_xml(xml_file_path)
    write_coordinates_to_excel(parsed_coordinates, excel_file_path)

    # Script 2: Load 3D medical image and create label map
    image_path = "march_9th_94fa1743/26001_GAD_STEALTH_3D_AX_T1_MPRAGE_1MM_C9C1HN021/26001_GAD_STEALTH_3D_AX_T1_MPRAGE_1MM_C9C1HN021.nii"
    image = sitk.ReadImage(image_path)

    excel_path = "_STEALTH_ROIs_.xlsx"
    coordinates_df = pd.read_excel(excel_path)
    physical_coordinates = coordinates_df[['x', 'y', 'z']].values

    label_map = sitk.Image(image.GetSize(), sitk.sitkUInt8)
    label_map.SetOrigin(image.GetOrigin())
    label_map.SetSpacing(image.GetSpacing())
    label_map.SetDirection(image.GetDirection())

    for i, coord in enumerate(physical_coordinates):
        index = image.TransformPhysicalPointToIndex(coord)
        label_map[index] = i + 1  # Indexing starts from 1

    label_map_path = "_STEALTH_ROIs_new.nii.gz"
    sitk.WriteImage(label_map, label_map_path)
